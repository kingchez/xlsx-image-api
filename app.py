from flask import Flask, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import Cell
from PIL import Image as PILImage
import requests
from io import BytesIO
import tempfile
import os
from urllib.parse import urlparse, urlunparse

app = Flask(__name__)

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        "status": "running",
        "message": "XLSX Image Creator API",
        "endpoints": {
            "/create-xlsx": "POST - Create XLSX with embedded images"
        }
    })

@app.route('/health', methods=['GET'])
def health():
    return jsonify({"status": "healthy"}), 200


def clean_image_url(url):
    """
    Strips CDN format-forcing parameters like f_avif, f_webp from URLs.
    e.g. CNN, Cloudinary, Imgix URLs that force a specific output format.
    This lets us download the original image instead of a converted version.
    """
    try:
        parsed = urlparse(url)
        # Remove f_* format directives from query string (e.g. f_avif, f_webp)
        clean_query = '&'.join(
            p for p in parsed.query.split('&')
            if not p.startswith('f_') and p != ''
        )
        return urlunparse(parsed._replace(query=clean_query))
    except Exception:
        return url  # if anything goes wrong, return original URL


def download_and_convert_image(url):
    """
    Downloads an image from any URL and converts it to PNG automatically.
    Handles ALL formats including:
      - Common:   .jpg, .jpeg, .png, .gif
      - Modern:   .webp, .avif, .heic
      - Other:    .bmp, .tiff, .ico, .svg (raster only)
      - CDN URLs: CNN, Cloudinary, Imgix with format parameters
    Returns a BytesIO PNG buffer, or None if anything fails.
    """
    try:
        # Step 1 — Clean the URL (strip format-forcing CDN params)
        clean_url = clean_image_url(url)
        if clean_url != url:
            print(f"Cleaned URL: {url} → {clean_url}")

        # Step 2 — Download with browser-like headers (some servers block bots)
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
        }

        img_response = requests.get(clean_url, timeout=15, headers=headers)

        # Step 3 — If cleaned URL fails, try the original URL as fallback
        if img_response.status_code != 200:
            print(f"Cleaned URL failed ({img_response.status_code}), trying original...")
            img_response = requests.get(url, timeout=15, headers=headers)

        if img_response.status_code != 200:
            print(f"Failed to download image: HTTP {img_response.status_code} — {url}")
            return None

        # Step 4 — Open with Pillow (auto-detects format, handles all types)
        img_data = BytesIO(img_response.content)
        pil_img = PILImage.open(img_data)

        print(f"Image format detected: {pil_img.format} | Mode: {pil_img.mode} | Size: {pil_img.size}")

        # Step 5 — Convert to RGB (handles transparency, palette, CMYK, etc.)
        if pil_img.mode == 'P':
            # Palette mode — convert to RGBA first to preserve any transparency
            pil_img = pil_img.convert('RGBA')

        if pil_img.mode in ('RGBA', 'LA'):
            # Has transparency — flatten onto white background
            background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
            background.paste(pil_img, mask=pil_img.split()[-1])
            pil_img = background
        elif pil_img.mode == 'CMYK':
            # CMYK (some JPEGs) — convert to RGB
            pil_img = pil_img.convert('RGB')
        elif pil_img.mode != 'RGB':
            # Any other mode (L, YCbCr, HSV, etc.) — convert to RGB
            pil_img = pil_img.convert('RGB')

        # Step 6 — Save as PNG into memory (openpyxl always gets a clean PNG)
        png_buffer = BytesIO()
        pil_img.save(png_buffer, format='PNG', optimize=True)
        png_buffer.seek(0)

        return png_buffer

    except Exception as e:
        print(f"Error processing image from {url}: {e}")
        return None


@app.route('/create-xlsx', methods=['POST'])
def create_xlsx():
    try:
        data = request.json
        items = data.get('items', [])

        if not items:
            return jsonify({"error": "No items provided"}), 400

        # Create workbook
        wb = Workbook(write_only=False)
        ws = wb.active
        ws.title = "Thumbnails"

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 40

        # Add headers row
        headers = ['ID', 'Category', 'Title Raw', 'Title 1', 'Title 2', 'Title 3', 'Design Type', 'Thumbnail', 'Name Base64']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.data_type = 's'

        # Add data rows
        for idx, item in enumerate(items, start=2):
            ws.cell(row=idx, column=1, value=item.get('id', ''))
            ws.cell(row=idx, column=2, value=item.get('category_name', ''))
            ws.cell(row=idx, column=3, value=item.get('title_raw', ''))
            ws.cell(row=idx, column=4, value=item.get('title1', ''))
            ws.cell(row=idx, column=5, value=item.get('title2', ''))
            ws.cell(row=idx, column=6, value=item.get('title3', ''))
            ws.cell(row=idx, column=7, value=item.get('design_type', ''))

            # Column H — placeholder empty cell
            empty_cell = ws.cell(row=idx, column=8, value=None)
            empty_cell.data_type = 'n'

            ws.cell(row=idx, column=9, value=item.get('name_base64', ''))

            # Download, auto-convert, and embed image
            thumbnail_url = item.get('thumbnail_url', '')
            if thumbnail_url:
                png_buffer = download_and_convert_image(thumbnail_url)
                if png_buffer:
                    try:
                        img = XLImage(png_buffer)
                        img.width = 150
                        img.height = 100
                        ws.add_image(img, f'H{idx}')
                        ws.row_dimensions[idx].height = 80
                    except Exception as e:
                        print(f"Error embedding image in row {idx}: {e}")
                else:
                    print(f"Skipping image for row {idx} — could not process: {thumbnail_url}")

        # Save to temp file and return
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            wb.save(temp_file.name)
            temp_path = temp_file.name

        response = send_file(
            temp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='thumbnails.xlsx'
        )

        @response.call_on_close
        def cleanup():
            try:
                os.unlink(temp_path)
            except:
                pass

        return response

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
